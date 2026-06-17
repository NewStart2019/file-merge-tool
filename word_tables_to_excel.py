#!/usr/bin/env python3
"""
Word表格提取到Excel工具
功能：
1. 保留Word表格样式、合并单元格（跨行跨列）
2. 按关键字过滤表头
3. 支持 .doc 转 .docx（LibreOffice）
4. 支持 PDF 转 Word 再提取
5. 多表格分Sheet存储
"""

import argparse
import os.path
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Optional

import office
import openpyxl
from docx import Document
from docx.oxml.ns import qn
from docx.table import Table
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ─────────────────────────── 颜色/样式工具 ───────────────────────────

def _docx_color_to_hex(color_obj) -> Optional[str]:
    """从 docx 颜色对象提取 hex 字符串（不含#），失败返回 None"""
    if color_obj is None:
        return None
    try:
        val = color_obj.rgb
        s = str(val)
        if s and s.upper() not in ("NONE", "AUTO"):
            return s
    except Exception:
        pass
    return None


def _xml_color_val(el, attr: str) -> Optional[str]:
    """从 XML 元素属性直接读颜色值"""
    if el is None:
        return None
    v = el.get(qn(attr)) or el.get(attr)
    if v and v.upper() not in ("AUTO", "NONE", ""):
        return v
    return None


def _make_side(border_elem) -> Side:
    """将 docx border element 转为 openpyxl Side。
    border_elem 是原始 lxml Element，属性用 qn 读取。"""
    if border_elem is None:
        return Side(border_style=None)
    from docx.oxml.ns import qn as _qn
    # 优先从 XML 属性直接读（兼容 lxml Element 和 docx 包装对象）
    val = border_elem.get(_qn("w:val")) or getattr(border_elem, "val", None)
    if val in (None, "none", "nil"):
        return Side(border_style=None)
    # 简单映射
    style_map = {
        "single": "thin",
        "thick": "thick",
        "double": "double",
        "dotted": "dotted",
        "dashed": "dashed",
        "dashDot": "dashDot",
        "dashDotDot": "dashDotDot",
        "medium": "medium",
        "mediumDashed": "mediumDashed",
        "mediumDashDot": "mediumDashDot",
    }
    xl_style = style_map.get(val, "thin")
    # 颜色：先从 XML 属性读，再尝试 docx 包装对象
    color = border_elem.get(_qn("w:color"))
    if not color or color.upper() == "AUTO":
        try:
            color = _docx_color_to_hex(border_elem.color)
        except Exception:
            color = None
    return Side(border_style=xl_style, color=color or "000000")


# ── 表格级默认样式提取 ──────────────────────────────────────────────────

class TableDefaults:
    """
    从 <w:tblPr> 提取表格全局默认值，供单元格回退使用。
    覆盖：边框、背景填充、单元格对齐、行高、字体（docStyle 继承）
    """
    def __init__(self, tbl_elem):
        from docx.oxml.ns import qn
        self.borders: dict[str, Optional[Side]] = {}   # top/bottom/left/right/insideH/insideV
        self.fill: Optional[str] = None
        self.valign: str = "center"
        self.halign: Optional[str] = None
        self.font_name: Optional[str] = None
        self.font_size: Optional[float] = None
        self.bold: bool = False
        self.italic: bool = False
        self.color: Optional[str] = None

        tblPr = tbl_elem.find(qn("w:tblPr"))
        if tblPr is None:
            return

        # 表格级边框 <w:tblBorders>
        tblBorders = tblPr.find(qn("w:tblBorders"))
        if tblBorders is not None:
            for tag in ("top", "bottom", "left", "right", "insideH", "insideV"):
                el = tblBorders.find(qn(f"w:{tag}"))
                self.borders[tag] = _make_side(el) if el is not None else None

        # 表格级背景 <w:shd>
        shd = tblPr.find(qn("w:shd"))
        if shd is not None:
            v = shd.get(qn("w:fill"))
            if v and v.upper() not in ("AUTO", "FFFFFF", "NONE"):
                self.fill = v

        # 表格水平对齐 <w:jc>（整体居中等）
        jc = tblPr.find(qn("w:jc"))
        if jc is not None:
            val = jc.get(qn("w:val"), "")
            mapping = {"left": "left", "center": "center", "right": "right",
                       "both": "justify", "start": "left", "end": "right"}
            self.halign = mapping.get(val)

        # 表格默认单元格属性 <w:tblCellPr>（不常见，兜底用）
        tblCellPr = tblPr.find(qn("w:tblCellPr"))
        if tblCellPr is not None:
            vAlign = tblCellPr.find(qn("w:vAlign"))
            if vAlign is not None:
                vmap = {"top": "top", "center": "center", "bottom": "bottom", "both": "center"}
                self.valign = vmap.get(vAlign.get(qn("w:val"), "center"), "center")

    def get_border_side(self, position: str, is_edge: bool) -> Side:
        """
        按位置返回应使用的边框 Side。
        position: top / bottom / left / right
        is_edge:  True = 表格外边框用 top/bottom/left/right
                  False = 内部格线用 insideH(上下) / insideV(左右)
        """
        # 外边框直接查对应方向
        if is_edge:
            s = self.borders.get(position)
            return s if s is not None else Side(border_style=None)
        # 内部格线：上下用 insideH，左右用 insideV
        inside_key = "insideH" if position in ("top", "bottom") else "insideV"
        s = self.borders.get(inside_key)
        if s is None:
            # 没有 inside 定义时回退到对应外边框
            s = self.borders.get(position)
        return s if s is not None else Side(border_style=None)


def _get_cell_border(tc, table_defaults: Optional["TableDefaults"] = None,
                     r: int = 0, c: int = 0,
                     num_rows: int = 1, max_cols: int = 1) -> Border:
    """
    获取单元格四边框，优先级：
      单元格自有 w:tcBorders > 行级 w:trPr（罕见）> 表格全局 tblBorders > 无边框
    外边框 vs 内部格线自动判断。
    """

    # ── 先读单元格自有边框 ────────────────────────────────────────────
    cell_sides: dict[str, Optional[Side]] = {}
    try:
        tcPr = tc.find(qn("w:tcPr"))
        if tcPr is not None:
            tcBorders = tcPr.find(qn("w:tcBorders"))
            if tcBorders is not None:
                for tag in ("top", "bottom", "left", "right"):
                    el = tcBorders.find(qn(f"w:{tag}"))
                    if el is not None:
                        cell_sides[tag] = _make_side(el)
    except Exception:
        pass

    def resolve(position: str) -> Side:
        # 1. 单元格自定义
        if position in cell_sides:
            return cell_sides[position]
        # 2. 表格全局回退
        if table_defaults is not None:
            is_edge = (
                (position == "top"    and r == 0) or
                (position == "bottom" and r == num_rows - 1) or
                (position == "left"   and c == 0) or
                (position == "right"  and c == max_cols - 1)
            )
            return table_defaults.get_border_side(position, is_edge)
        return Side(border_style=None)

    return Border(
        top=resolve("top"), bottom=resolve("bottom"),
        left=resolve("left"), right=resolve("right"),
    )


def _get_cell_fill(tc, table_defaults: Optional["TableDefaults"] = None) -> Optional[str]:
    """从 docx tc XML 获取背景填充色 hex，无填充时回退到表格全局默认"""
    try:
        tcPr = tc.find(qn("w:tcPr"))
        if tcPr is not None:
            shd = tcPr.find(qn("w:shd"))
            if shd is not None:
                fill_val = shd.get(qn("w:fill"))
                if fill_val and fill_val.upper() not in ("AUTO", "FFFFFF", "NONE"):
                    return fill_val
    except Exception:
        pass
    # 回退：表格全局背景
    if table_defaults is not None and table_defaults.fill:
        return table_defaults.fill
    return None


def _get_cell_valign(tc, table_defaults: Optional["TableDefaults"] = None) -> str:
    """获取单元格垂直对齐，无则回退到表格全局默认"""
    try:
        tcPr = tc.find(qn("w:tcPr"))
        if tcPr is not None:
            vAlign = tcPr.find(qn("w:vAlign"))
            if vAlign is not None:
                val = vAlign.get(qn("w:val"), "")
                mapping = {"top": "top", "center": "center", "bottom": "bottom", "both": "center"}
                result = mapping.get(val)
                if result:
                    return result
    except Exception:
        pass
    # 回退：表格全局
    if table_defaults is not None:
        return table_defaults.valign
    return "center"


# ─────────────────────────── Run 级样式提取 ───────────────────────────

class RunStyle:
    """单个 run 的完整字体样式"""
    __slots__ = ("text", "bold", "italic", "underline", "strike",
                 "font_name", "font_size", "color", "halign")

    def __init__(self):
        self.text = ""
        self.bold = False
        self.italic = False
        self.underline = False
        self.strike = False
        self.font_name: Optional[str] = None
        self.font_size: Optional[float] = None   # pt
        self.color: Optional[str] = None         # hex
        self.halign: Optional[str] = None        # left/center/right/justify


def _extract_run_style(rPr, table_defaults: Optional["TableDefaults"] = None) -> dict:
    """
    从 <w:rPr> XML 元素提取完整字体属性，返回字典。
    table_defaults: 表格全局默认（字体名/字号/颜色回退）
    """
    # 预填表格级默认值（作为 fallback）
    td = table_defaults
    props = {
        "bold":      td.bold      if td else False,
        "italic":    td.italic    if td else False,
        "underline": False,
        "strike":    False,
        "font_name": td.font_name if td else None,
        "font_size": td.font_size if td else None,
        "color":     td.color     if td else None,
    }
    if rPr is None:
        return props

    # 粗体：<w:b> 存在且 w:val != "0"
    b_el = rPr.find(qn("w:b"))
    if b_el is not None:
        props["bold"] = b_el.get(qn("w:val"), "1") not in ("0", "false")

    # 斜体
    i_el = rPr.find(qn("w:i"))
    if i_el is not None:
        props["italic"] = i_el.get(qn("w:val"), "1") not in ("0", "false")

    # 下划线
    u_el = rPr.find(qn("w:u"))
    if u_el is not None:
        uval = u_el.get(qn("w:val"), "")
        props["underline"] = uval not in ("none", "0", "")

    # 删除线
    strike_el = rPr.find(qn("w:strike"))
    if strike_el is not None:
        props["strike"] = strike_el.get(qn("w:val"), "1") not in ("0", "false")

    # 字体名称：优先 w:eastAsia（中文），次选 w:ascii
    fonts_el = rPr.find(qn("w:rFonts"))
    if fonts_el is not None:
        ea = fonts_el.get(qn("w:eastAsia"))
        ascii_ = fonts_el.get(qn("w:ascii"))
        props["font_name"] = ea or ascii_

    # 字号：<w:sz> 单位是半磅，除以 2 得 pt
    sz_el = rPr.find(qn("w:sz"))
    if sz_el is not None:
        try:
            props["font_size"] = int(sz_el.get(qn("w:val"), 0)) / 2
        except (ValueError, TypeError):
            pass

    # 字体颜色
    color_el = rPr.find(qn("w:color"))
    if color_el is not None:
        v = color_el.get(qn("w:val"), "")
        if v and v.upper() not in ("AUTO", ""):
            props["color"] = v

    return props


def _get_para_halign(p) -> Optional[str]:
    """从 <w:p> 获取水平对齐方式"""
    try:
        pPr = p.find(qn("w:pPr"))
        if pPr is None:
            return None
        jc = pPr.find(qn("w:jc"))
        if jc is None:
            return None
        val = jc.get(qn("w:val"), "")
        mapping = {
            "left": "left", "start": "left",
            "center": "center",
            "right": "right", "end": "right",
            "both": "justify", "distribute": "justify",
        }
        return mapping.get(val)
    except Exception:
        return None


def _merge_run_styles(styles: list[dict]) -> dict:
    """
    合并多个 run 的样式：
    - 字体名/大小/颜色取第一个非空值
    - 粗体/斜体/下划线：任一为 True 则为 True
    """
    merged = {
        "bold": False, "italic": False, "underline": False, "strike": False,
        "font_name": None, "font_size": None, "color": None,
    }
    for s in styles:
        merged["bold"] = merged["bold"] or s.get("bold", False)
        merged["italic"] = merged["italic"] or s.get("italic", False)
        merged["underline"] = merged["underline"] or s.get("underline", False)
        merged["strike"] = merged["strike"] or s.get("strike", False)
        if merged["font_name"] is None:
            merged["font_name"] = s.get("font_name")
        if merged["font_size"] is None:
            merged["font_size"] = s.get("font_size")
        if merged["color"] is None:
            merged["color"] = s.get("color")
    return merged


# ─────────────────────────── 合并单元格解析 ───────────────────────────

def _parse_merge_map(table: Table):
    """
    返回 merge_map: dict[(row_idx, col_idx)] -> (rowspan, colspan)
    以及 skip_set: 被合并覆盖、不需要写入的单元格集合
    """
    num_rows = len(table.rows)
    # 先收集每行列数
    grid = []
    for row in table.rows:
        row_cells = []
        for tc in row._tr.findall(qn("w:tc")):
            tcPr = tc.find(qn("w:tcPr"))
            colspan = 1
            if tcPr is not None:
                gridSpan = tcPr.find(qn("w:gridSpan"))
                if gridSpan is not None:
                    colspan = int(gridSpan.get(qn("w:val"), 1))
            row_cells.append((tc, colspan))
        grid.append(row_cells)

    # 构建完整网格（处理 vMerge）
    # grid_full[r][c] = (tc, is_first_of_vmerge)
    max_cols = sum(c for _, c in grid[0]) if grid else 0
    grid_full = [[None] * max_cols for _ in range(num_rows)]

    # 先填 colspan
    for r, row_cells in enumerate(grid):
        c = 0
        for tc, colspan in row_cells:
            for dc in range(colspan):
                if c + dc < max_cols:
                    grid_full[r][c + dc] = tc
            c += colspan

    # 处理 vMerge → 计算 rowspan
    merge_map = {}   # (r,c) -> (rowspan, colspan)
    skip_set = set()

    def get_gridspan(tc):
        tcPr = tc.find(qn("w:tcPr"))
        if tcPr is None:
            return 1
        gs = tcPr.find(qn("w:gridSpan"))
        return int(gs.get(qn("w:val"), 1)) if gs is not None else 1

    def is_vmerge_restart(tc):
        tcPr = tc.find(qn("w:tcPr"))
        if tcPr is None:
            return False
        vm = tcPr.find(qn("w:vMerge"))
        if vm is None:
            return False
        val = vm.get(qn("w:val"), "")
        return val == "restart"

    def is_vmerge_continue(tc):
        tcPr = tc.find(qn("w:tcPr"))
        if tcPr is None:
            return False
        vm = tcPr.find(qn("w:vMerge"))
        if vm is None:
            return False
        val = vm.get(qn("w:val"), "")
        return val != "restart"

    visited = set()
    for r in range(num_rows):
        c = 0
        while c < max_cols:
            tc = grid_full[r][c]
            if tc is None or (r, c) in visited:
                c += 1
                continue

            colspan = get_gridspan(tc)

            if is_vmerge_restart(tc):
                # 计算 rowspan
                rowspan = 1
                for rr in range(r + 1, num_rows):
                    next_tc = grid_full[rr][c]
                    if next_tc is not None and is_vmerge_continue(next_tc):
                        rowspan += 1
                        # 标记 continue 行的所有列为 skip
                        for dc in range(colspan):
                            skip_set.add((rr, c + dc))
                            visited.add((rr, c + dc))
                    else:
                        break
                merge_map[(r, c)] = (rowspan, colspan)
                for dc in range(1, colspan):
                    skip_set.add((r, c + dc))
            elif is_vmerge_continue(tc):
                skip_set.add((r, c))
            else:
                if colspan > 1:
                    merge_map[(r, c)] = (1, colspan)
                    for dc in range(1, colspan):
                        skip_set.add((r, c + dc))

            visited.add((r, c))
            c += colspan

    return merge_map, skip_set


# ─────────────────────────── 表格内容提取 ───────────────────────────

def _get_table_header_text(table: Table) -> str:
    """获取表格第一行所有文字拼接"""
    try:
        first_row = table.rows[0]
        texts = [cell.text.strip() for cell in first_row.cells]
        return " ".join(texts)
    except Exception:
        return ""


def _get_table_all_text(table: Table) -> str:
    """获取表格全部文字"""
    parts = []
    for row in table.rows:
        for cell in row.cells:
            parts.append(cell.text.strip())
    return " ".join(parts)


def _table_matches_keywords(table: Table, keywords: list[str], search_all: bool = False) -> bool:
    """
    检查表格是否包含关键字。
    search_all=False 只检查第一行（表头），=True 检查全部内容。
    """
    if not keywords:
        return True
    text = _get_table_all_text(table) if search_all else _get_table_header_text(table)
    text_lower = text.lower()
    return any(kw.lower() in text_lower for kw in keywords)


# ─────────────────────────── 写入 Excel ───────────────────────────

def _write_table_to_sheet(ws, table: Table, merge_map: dict, skip_set: set):
    """将一张 docx 表格写入 openpyxl worksheet，完整保留样式和合并"""
    num_rows = len(table.rows)

    # ── 提取表格全局默认样式（边框/填充/对齐的回退来源）────────────────
    tbl_defaults = TableDefaults(table._tbl)

    # ── 展开网格 ──────────────────────────────────────────────────────
    grid_tcs = []
    for row in table.rows:
        row_tcs = []
        for tc in row._tr.findall(qn("w:tc")):
            tcPr = tc.find(qn("w:tcPr"))
            colspan = 1
            if tcPr is not None:
                gs = tcPr.find(qn("w:gridSpan"))
                if gs is not None:
                    colspan = int(gs.get(qn("w:val"), 1))
            row_tcs.append((tc, colspan))
        grid_tcs.append(row_tcs)

    max_cols = sum(c for _, c in grid_tcs[0]) if grid_tcs else 0

    # 展开成完整网格
    full_tc = [[None] * max_cols for _ in range(num_rows)]
    for r, row_tcs in enumerate(grid_tcs):
        c = 0
        for tc, colspan in row_tcs:
            for dc in range(colspan):
                if c + dc < max_cols:
                    full_tc[r][c + dc] = tc
            c += colspan

    # ── 逐单元格写入 ──────────────────────────────────────────────────
    written_tcs = set()

    for r in range(num_rows):
        xl_row = r + 1
        for c in range(max_cols):
            xl_col = c + 1

            if (r, c) in skip_set:
                continue

            tc = full_tc[r][c]
            if tc is None:
                continue

            # 防重复（同一 tc 跨 colspan 只取第一次）
            tc_id = id(tc)
            if tc_id in written_tcs and (r, c) not in merge_map:
                continue
            written_tcs.add(tc_id)

            # 文本
            cell_obj = ws.cell(row=xl_row, column=xl_col)
            paragraphs = tc.findall(qn("w:p"))

            # ── 文本 + 每段样式收集 ───────────────────────────────────
            lines = []
            all_run_styles = []   # 所有 run 的样式列表
            para_haligns = []     # 每段的水平对齐

            for p in paragraphs:
                line_parts = []
                para_haligns.append(_get_para_halign(p))

                for run in p.findall(qn("w:r")):
                    # 跳过删除标记内的 run
                    if run.getparent().tag == qn("w:del"):
                        continue
                    t = run.find(qn("w:t"))
                    if t is not None and t.text:
                        line_parts.append(t.text)
                    # 取第一个有颜色/粗体的 run
                    rPr = run.find(qn("w:rPr"))
                    style = _extract_run_style(rPr, tbl_defaults)
                    if line_parts:  # 只收集有文字的 run 的样式
                        all_run_styles.append(style)

                lines.append("".join(line_parts))

            cell_obj.value = "\n".join(lines)

            # ── 合并所有 run 的样式，取主样式 ────────────────────────
            merged = _merge_run_styles(all_run_styles) if all_run_styles else {
                "bold": False, "italic": False, "underline": False,
                "strike": False, "font_name": None, "font_size": None, "color": None,
            }

            # ── 字体 ─────────────────────────────────────────────────
            font_kwargs = {
                "bold":   merged["bold"],
                "italic": merged["italic"],
                "strike": merged["strike"],
                "name":   merged["font_name"] or "Arial",
                "size":   merged["font_size"] or 10,
            }
            # openpyxl underline 需要字符串，不能传 bool
            if merged["underline"]:
                font_kwargs["underline"] = "single"
            if merged["color"]:
                font_kwargs["color"] = merged["color"]
            cell_obj.font = Font(**font_kwargs)

            # ── 水平对齐（取段落中出现最多的对齐方式，无则回退表格默认）──
            halign = None
            valid_aligns = [a for a in para_haligns if a]
            if valid_aligns:
                halign = max(set(valid_aligns), key=valid_aligns.count)
            if halign is None and tbl_defaults.halign:
                halign = tbl_defaults.halign

            # ── 垂直对齐 ─────────────────────────────────────────────
            valign = _get_cell_valign(tc, tbl_defaults)

            cell_obj.alignment = Alignment(
                horizontal=halign,       # None = Excel 默认（左对齐）
                vertical=valign,
                wrap_text=True,
            )

            # ── 背景填充 ─────────────────────────────────────────────
            fill_hex = _get_cell_fill(tc, tbl_defaults)
            if fill_hex:
                cell_obj.fill = PatternFill("solid", fgColor=fill_hex)

            # ── 边框 ─────────────────────────────────────────────────
            cell_obj.border = _get_cell_border(
                tc, tbl_defaults, r, c, num_rows, max_cols)

            # ── 行高（从段落间距推算，最低 15pt）────────────────────
            try:
                pPr = paragraphs[0].find(qn("w:pPr")) if paragraphs else None
                if pPr is not None:
                    spacing = pPr.find(qn("w:spacing"))
                    if spacing is not None:
                        before = int(spacing.get(qn("w:before"), 0) or 0)
                        after  = int(spacing.get(qn("w:after"),  0) or 0)
                        line   = int(spacing.get(qn("w:line"),   0) or 0)
                        if line > 0:
                            # line 单位是 twip/20，转 pt
                            row_h = max(line / 20, 15)
                            ws.row_dimensions[xl_row].height = row_h
            except Exception:
                pass

            # ── 合并单元格 ───────────────────────────────────────────
            if (r, c) in merge_map:
                rowspan, colspan = merge_map[(r, c)]
                if rowspan > 1 or colspan > 1:
                    ws.merge_cells(
                        start_row=xl_row, start_column=xl_col,
                        end_row=xl_row + rowspan - 1,
                        end_column=xl_col + colspan - 1,
                    )

    # ── 自动列宽（按字符数估算，中文字符算 2 宽）─────────────────────
    for col in ws.columns:
        max_w = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                for ln in val.split("\n"):
                    # 中文/全角字符宽度按 2 计
                    w = sum(2 if ord(ch) > 127 else 1 for ch in ln)
                    if w > max_w:
                        max_w = w
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_w + 2, 8), 60)


# ─────────────────────────── 格式转换 ───────────────────────────

def convert_doc_to_docx(doc_path: str, out_dir: Optional[str] = None) -> str:
    """
    .doc → .docx，优先 LibreOffice，备选 python-office。
    out_dir: 输出目录，None 则与源文件同目录。
    返回生成的 .docx 路径。
    """
    doc_path = str(doc_path)
    out_dir = str(out_dir) if out_dir else str(Path(doc_path).parent)
    Path(out_dir).mkdir(parents=True, exist_ok=True)

    # 方法1：LibreOffice（最可靠）
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if soffice:
        result = subprocess.run(
            [soffice, "--headless", "--convert-to", "docx",
             "--outdir", out_dir, doc_path],
            capture_output=True, text=True, timeout=120
        )
        docx_path = str(Path(out_dir) / (Path(doc_path).stem + ".docx"))
        if Path(docx_path).exists():
            print(f"  [LibreOffice] doc → docx: {docx_path}")
            return docx_path
        print(f"  [LibreOffice stderr] {result.stderr}")

    # 方法2：python-office
    try:
        docx_path = str(Path(out_dir) / (Path(doc_path).stem + ".docx"))
        office.word.doc2docx(doc_path,os.path.dirname(docx_path), os.path.basename(docx_path))
        if Path(docx_path).exists():
            print(f"  [python-office] doc → docx: {docx_path}")
            return docx_path
    except Exception as e:
        print(f"  [python-office] 转换失败: {e}")

    raise RuntimeError(
        f"无法将 {doc_path} 转为 docx，请安装 LibreOffice 或 python-office。"
    )


def convert_pdf_to_docx(pdf_path: str, out_dir: Optional[str] = None) -> str:
    """
    PDF → .docx（pdf2docx），返回生成的 .docx 路径。
    out_dir: 输出目录，None 则与源文件同目录。
    """
    from pdf2docx import Converter
    out_dir = str(out_dir) if out_dir else str(Path(pdf_path).parent)
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    docx_path = str(Path(out_dir) / (Path(pdf_path).stem + ".docx"))
    print(f"  [pdf2docx] 正在转换 PDF → docx，可能需要一点时间…")
    cv = Converter(pdf_path)
    cv.convert(docx_path, start=0, end=None)
    cv.close()
    print(f"  [pdf2docx] 完成: {docx_path}")
    return docx_path


# ─────────────────────────── 主提取函数 ───────────────────────────

def extract_tables_to_excel(
    input_path: str,
    output_path: str,
    keywords: Optional[list[str]] = None,
    search_header_only: bool = True,
    data_dir: Optional[str] = None,
):
    """
    从 Word（.docx / .doc）或 PDF 中提取表格到 Excel。

    参数:
        input_path       : 输入文件路径（.docx / .doc / .pdf）
        output_path      : 输出 .xlsx 路径
        keywords         : 表头关键字列表，None 表示提取全部表格
        search_header_only: True=只检查表头行，False=检查全部内容
        data_dir         : 中间转换文件的存放目录，None 则与输入文件同目录
    """
    input_path = str(input_path)
    suffix = Path(input_path).suffix.lower()

    # 格式转换
    docx_path = input_path
    if suffix == ".doc":
        print(f"[步骤1] 检测到 .doc，转换为 .docx …")
        docx_path = convert_doc_to_docx(input_path, out_dir=data_dir)
    elif suffix == ".pdf":
        print(f"[步骤1] 检测到 .pdf，转换为 .docx …")
        docx_path = convert_pdf_to_docx(input_path, out_dir=data_dir)
    else:
        print(f"[步骤1] 输入文件: {input_path}")

    # 读取 Word 文档
    print(f"[步骤2] 读取 Word 文档: {docx_path}")
    doc = Document(docx_path)
    all_tables = doc.tables
    print(f"  发现 {len(all_tables)} 张表格")

    # 关键字过滤
    if keywords:
        selected = [
            (i, t) for i, t in enumerate(all_tables)
            if _table_matches_keywords(t, keywords, not search_header_only)
        ]
        print(f"  关键字过滤后剩余: {len(selected)} 张表格 (关键字={keywords})")
    else:
        selected = list(enumerate(all_tables))

    if not selected:
        print("  未找到符合条件的表格，退出。")
        return

    # 创建 Excel
    print(f"[步骤3] 写入 Excel: {output_path}")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    for sheet_idx, (orig_idx, table) in enumerate(selected):
        sheet_name = f"表格{sheet_idx + 1}"
        ws = wb.create_sheet(title=sheet_name)

        # 解析合并信息
        merge_map, skip_set = _parse_merge_map(table)

        # 写入
        _write_table_to_sheet(ws, table, merge_map, skip_set)

        rows = len(table.rows)
        cols = len(table.columns)
        header_text = _get_table_header_text(table)[:60]
        print(f"  Sheet [{sheet_name}] ← 原始表格#{orig_idx + 1} "
              f"({rows}行×{cols}列) 表头: {header_text!r}")

    if output_path is None or output_path == "":
        output_path = os.path.join(data_dir, os.path.basename(input_path))
    wb.save(output_path)
    print(f"\n✅ 完成！共导出 {len(selected)} 张表格 → {output_path}")


# ─────────────────────────── 批量扫描 ───────────────────────────

def scan_and_extract(
    scan_dir: str,
    keywords: Optional[list[str]] = None,
    search_header_only: bool = True,
):
    """
    扫描目录下所有 .docx / .doc / .pdf，逐个提取表格到 Excel。

    目录结构：
        scan_dir/
        ├── data/                ← 自动创建
        │   ├── xxx.docx         ← PDF/DOC 转换的中间文件
        │   └── xxx_tables.xlsx  ← 最终 Excel 输出
        └── *.docx / *.pdf / *.doc  ← 原始输入文件（只读，不修改）
    """
    scan_dir  = Path(scan_dir).resolve()
    # 单文件模式：中间文件和 excel 都写到 当前目录/data/
    data_dir = Path.cwd() / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    # 收集所有支持的文件（跳过 data 子目录，避免重复处理中间文件）
    suffixes = {".docx", ".doc", ".pdf"}
    files = sorted([
        p for p in scan_dir.iterdir()
        if p.is_file() and p.suffix.lower() in suffixes
    ])

    if not files:
        print(f"⚠️  目录下未找到 .docx / .doc / .pdf 文件：{scan_dir}")
        return

    total = len(files)
    print(f"\n📂 扫描目录: {scan_dir}")
    print(f"   找到 {total} 个文件，输出目录: {data_dir}")
    print("=" * 60)

    ok_list   = []   # 成功
    skip_list = []   # 无表格
    fail_list = []   # 报错

    for idx, file_path in enumerate(files, 1):
        stem        = file_path.stem
        output_path = str(data_dir / f"{stem}_tables.xlsx")

        print(f"\n[{idx}/{total}] {file_path.name}")
        print(f"  → {output_path}")

        try:
            extract_tables_to_excel(
                input_path=str(file_path),
                output_path=output_path,
                keywords=keywords,
                search_header_only=search_header_only,
                data_dir=str(data_dir),
            )
            if Path(output_path).exists():
                ok_list.append(file_path.name)
            else:
                skip_list.append(file_path.name)
        except Exception as e:
            print(f"  ❌ 处理失败: {e}")
            fail_list.append((file_path.name, str(e)))

    # 汇总报告
    print("\n" + "=" * 60)
    print(f"📊 批量处理完成")
    print(f"   ✅ 成功: {len(ok_list)} 个")
    if ok_list:
        for name in ok_list:
            print(f"      · {name}")
    if skip_list:
        print(f"   ⚠️  无表格(已跳过): {len(skip_list)} 个")
        for name in skip_list:
            print(f"      · {name}")
    if fail_list:
        print(f"   ❌ 失败: {len(fail_list)} 个")
        for name, err in fail_list:
            print(f"      · {name}  原因: {err}")
    print(f"\n   输出目录: {data_dir}")


# ─────────────────────────── CLI ───────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Word/PDF 表格提取到 Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # ── 单文件模式 ──────────────────────────────────────────
  # 提取全部表格
  python word_tables_to_excel.py report.docx

  # 只提取表头含"合同"或"金额"的表格
  python word_tables_to_excel.py report.docx -k 合同 金额

  # 指定输出路径
  python word_tables_to_excel.py report.docx out.xlsx

  # PDF / DOC 文件（自动转换）
  python word_tables_to_excel.py report.pdf -k 姓名

  # 全文搜索关键字
  python word_tables_to_excel.py report.docx -k 数据 --search-all

  # ── 批量目录模式 ─────────────────────────────────────────
  # 扫描目录下所有 docx/pdf，结果写到 目录/data/
  python word_tables_to_excel.py -d ./documents

  # 批量 + 关键字过滤
  python word_tables_to_excel.py -d ./documents -k 合同 金额
"""
    )

    # 互斥：单文件 or 目录批量
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("input", nargs="?",
                       help="单个输入文件 (.docx / .doc / .pdf)")
    group.add_argument("-d", "--dir", metavar="DIR",
                       help="批量模式：扫描该目录下所有 docx/doc/pdf 文件")

    parser.add_argument("output", nargs="?",
                        help="单文件模式：输出 Excel 路径（默认自动生成）")
    parser.add_argument("-k", "--keywords", nargs="+", metavar="KW",
                        help="表头关键字（任意一个匹配即保留，不填则提取全部）")
    parser.add_argument("--search-all", action="store_true",
                        help="在全表内容中搜索关键字（默认只搜表头第一行）")

    args = parser.parse_args()

    kw             = args.keywords
    header_only    = not args.search_all

    # ── 批量目录模式 ──────────────────────────────────────────────────
    if args.dir:
        scan_dir = args.dir
        if not Path(scan_dir).is_dir():
            print(f"错误：目录不存在：{scan_dir}")
            sys.exit(1)
        scan_and_extract(
            scan_dir=scan_dir,
            keywords=kw,
            search_header_only=header_only,
        )
        return

    # ── 单文件模式 ────────────────────────────────────────────────────
    input_path = args.input
    if not input_path:
        parser.print_help()
        sys.exit(1)
    if not Path(input_path).exists():
        print(f"错误：文件不存在：{input_path}")
        sys.exit(1)

    # 单文件模式：中间文件和 excel 都写到 当前目录/data/
    data_dir = Path.cwd() / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    if args.output:
        output_path = args.output
    else:
        stem = Path(input_path).stem
        output_path = str(data_dir / f"{stem}_tables.xlsx")

    extract_tables_to_excel(
        input_path=input_path,
        output_path=output_path,
        keywords=kw,
        search_header_only=header_only,
        data_dir=str(data_dir),
    )


if __name__ == "__main__":
    main()